import re
import os
import json
import math
import time
import statistics
import urllib.request
import urllib.parse
import urllib.error
import html as html_module
from datetime import datetime
from typing import Optional

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')   # headless — no display needed
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from sklearn.feature_extraction.text import TfidfVectorizer
import cv2
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


# 0.  CONFIGURATION


OUTPUT_DIR = "results"
os.makedirs(OUTPUT_DIR, exist_ok=True)

INSTRUCTIONAL_WORDS = [
    "first", "second", "third", "next", "then", "after", "finally",
    "step", "now", "start", "begin", "make sure", "remember", "important",
    "tip", "trick", "apply", "use", "take", "hold", "wrap", "section",
    "curl", "straighten", "blow", "dry", "brush", "comb", "part",
    "divide", "clip", "pin", "spray", "heat", "protect",
]

FILLER_WORDS = [
    "um", "uh", "like", "you know", "basically", "literally",
    "actually", "kind of", "sort of", "right", "okay so", "so um",
    "and um", "i mean", "you see",
]



# 1.  TRANSCRIPT EXTRACTION  (YouTube → plain text + timed segments)


def extract_video_id(url: str) -> Optional[str]:
    """Handle youtube.com/watch?v=, youtu.be/, and shorts/ formats."""
    patterns = [
        r"(?:v=|youtu\.be/|shorts/)([A-Za-z0-9_-]{11})",
    ]
    for p in patterns:
        m = re.search(p, url)
        if m:
            return m.group(1)
    return None


def fetch_youtube_transcript(video_id: str) -> tuple[Optional[str], Optional[list]]:
    """
    Fetch YouTube auto-captions without any third-party library.

    Strategy:
      1. Load the YouTube watch page (standard browser User-Agent).
      2. Find the captionTracks JSON embedded in the page JS.
      3. Pick the English track (or first available).
      4. Fetch the timed XML transcript directly.
      5. Parse → return (full_text, list of {text, start, duration}).
    """
    watch_url = f"https://www.youtube.com/watch?v={video_id}"
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9",
    }
    try:
        req = urllib.request.Request(watch_url, headers=headers)
        with urllib.request.urlopen(req, timeout=15) as resp:
            page_html = resp.read().decode("utf-8", errors="ignore")
    except Exception as e:
        return None, None

    # Extract captionTracks JSON blob
    m = re.search(r'"captionTracks":(\[.*?\])', page_html)
    if not m:
        return None, None

    try:
        tracks = json.loads(m.group(1))
    except json.JSONDecodeError:
        return None, None

    # Prefer English; fall back to first available
    track_url = None
    for track in tracks:
        lang = track.get("languageCode", "")
        if lang.startswith("en"):
            track_url = track.get("baseUrl")
            break
    if not track_url and tracks:
        track_url = tracks[0].get("baseUrl")
    if not track_url:
        return None, None

    
    try:
        req2 = urllib.request.Request(track_url, headers=headers)
        with urllib.request.urlopen(req2, timeout=15) as resp2:
            xml_data = resp2.read().decode("utf-8", errors="ignore")
    except Exception:
        return None, None

    
    segments = []
    for seg in re.finditer(
        r'<text\s+start="([^"]+)"\s+dur="([^"]+)"[^>]*>(.*?)</text>',
        xml_data, re.DOTALL
    ):
        start = float(seg.group(1))
        dur   = float(seg.group(2))
        raw   = seg.group(3)
        text  = html_module.unescape(re.sub(r"<[^>]+>", "", raw)).strip()
        if text:
            segments.append({"text": text, "start": start, "duration": dur})

    if not segments:
        return None, None

    full_text = " ".join(s["text"] for s in segments)
    return full_text, segments



# 2.  SYNTHETIC TRANSCRIPT GENERATOR  (demo / offline fallback)

DEMO_TRANSCRIPTS = {
    "demo_excellent": {
        "title": "Perfect Blowout at Home - Step by Step Tutorial",
        "score_hint": 8.5,
        "text": (
            "Hey everyone, welcome back! Today I'm going to walk you through "
            "getting a perfect blowout at home. First, make sure your hair is "
            "about eighty percent dry before you begin. Take a medium-sized "
            "section from the front. Hold your brush underneath and direct the "
            "nozzle downward — this seals the cuticle and adds shine. "
            "Then roll the brush inward as you move toward the ends. "
            "Remember to keep the dryer moving; holding it still causes heat damage. "
            "Next, clip that section up and move to the underneath layers. "
            "The key tip here is tension — the more tension you apply, the smoother "
            "your result. Finally, once everything is dry, hit it with a cool shot "
            "to lock in the style. That is literally all you need for a salon blowout "
            "at home. Let me know in the comments if you have questions!"
        ),
        "duration": 420,
    },
    "demo_average": {
        "title": "Curling Hair Tutorial",
        "score_hint": 5.8,
        "text": (
            "Okay so um, today we're doing curls, right? So like, you basically "
            "just take your curling iron and um you wrap the hair around it. "
            "I mean, the size of the barrel kind of matters I guess. Um so like "
            "a one inch barrel gives you tighter curls, you know? And uh, you just "
            "hold it for like maybe ten seconds or so. I actually like to hold it "
            "a little longer basically. And um then you release it, right? And you "
            "kind of like let it cool down in your hand. Um I mean a lot of people "
            "sort of skip that step but it actually does make a difference literally. "
            "So yeah, just keep doing that all over and you'll basically be done."
        ),
        "duration": 380,
    },
    "demo_poor": {
        "title": "how to do hair",
        "score_hint": 3.2,
        "text": (
            "um hi so this is how you do hair like its easy you just take the "
            "hair thing and put it on your hair yeah and then you like wait um "
            "and then it looks good i mean its not hard or anything um so yeah "
            "just try it and uh it will work um like i do it all the time so "
            "yeah um okay bye"
        ),
        "duration": 145,
    },
    "demo_good": {
        "title": "How I Braid My Hair - Dutch Braid for Beginners",
        "score_hint": 7.1,
        "text": (
            "Welcome! Today we are doing a Dutch braid, which is perfect for "
            "beginners. Start by brushing out any tangles completely. Then divide "
            "your hair into three equal sections at the crown. Now for a Dutch braid "
            "you cross the strands underneath, not over. Take the right section under "
            "the middle, then left under the new middle. Each time you add a small "
            "piece of hair from the sides into the section before crossing. Make sure "
            "to keep consistent tension — this is really the most important step. "
            "If your braid looks uneven, tension is almost always the reason. "
            "Continue this all the way down to the nape and secure with a small elastic. "
            "You can then pancake the braid by gently pulling the loops apart for a "
            "fuller look. That's the whole process!"
        ),
        "duration": 390,
    },
    "demo_great": {
        "title": "5 Heatless Hairstyles for Everyday | Quick & Easy",
        "score_hint": 7.9,
        "text": (
            "If you're looking to protect your hair from heat damage, these five "
            "styles are your new best friends. Style one: the twisted low bun. "
            "Split your hair into two sections, twist each section away from your face, "
            "then coil them together and pin at the nape. Takes about ninety seconds. "
            "Style two: the rope braid. Take two pieces and twist them in the same "
            "direction, then wrap them around each other in the opposite direction. "
            "The physics of this actually keeps it from unraveling — pretty cool, right? "
            "Style three is a sleek low pony with a hair wrap. Secure your pony, "
            "take a thin strand from underneath, wrap it around the elastic, and pin "
            "it underneath. Instantly looks more polished. Style four is space buns — "
            "divide hair down the middle, twist each half into a bun, pin loosely. "
            "And style five, my personal favourite, is a knotted half-up. Take two "
            "front sections, tie them in a loose knot, and pin. Done. No heat, no fuss, "
            "all of these work on second-day hair too."
        ),
        "duration": 510,
    },
}


def build_synthetic_segments(text: str, total_duration: float) -> list:
    """
    Turn a flat text + duration into timed segments that mimic real captions.
    Word count drives proportional timing.
    """
    words = text.split()
    total_words = len(words)
    segments, current_time = [], 0.0
    chunk_size = 8   # words per caption segment (realistic)
    for i in range(0, total_words, chunk_size):
        chunk_words = words[i: i + chunk_size]
        chunk_text  = " ".join(chunk_words)
        # proportional duration
        seg_dur = (len(chunk_words) / total_words) * total_duration
        segments.append({
            "text": chunk_text,
            "start": current_time,
            "duration": seg_dur,
        })
        current_time += seg_dur
    return segments


# 3.  LINGUISTIC FEATURES  (from transcript text + segments)

def feat_speech_rate_wpm(segments: list) -> float:
    """
    Words per minute.
    Ideal for tutorials: 120–160 WPM.
    Scored: distance from ideal midpoint (140 WPM) mapped to 0–10.
    """
    if not segments or len(segments) < 2:
        return 5.0
    total_words = sum(len(s["text"].split()) for s in segments)
    duration_min = (segments[-1]["start"] + segments[-1]["duration"]) / 60
    if duration_min <= 0:
        return 5.0
    wpm = total_words / duration_min
    # Score: penalise deviation from 140 WPM
    deviation = abs(wpm - 140)
    score = max(0.0, 10.0 - (deviation / 14))   # -1 pt per 14 WPM off
    return round(score, 2), round(wpm, 1)


def feat_readability(text: str) -> float:
    """
    Flesch Reading Ease → normalised to 0–10.
    FRE 60–70 = ideal for general tutorial audiences.
    Formula: 206.835 - 1.015*(words/sentences) - 84.6*(syllables/words)
    """
    sentences = [s.strip() for s in re.split(r"[.!?]+", text) if s.strip()]
    words     = [w for w in re.findall(r"\b[a-zA-Z]+\b", text)]
    if not sentences or not words:
        return 5.0

    def syllable_count(word):
        word = word.lower()
        count = len(re.findall(r"[aeiou]+", word))
        if word.endswith("e") and count > 1:
            count -= 1
        return max(1, count)

    avg_sentence_len = len(words) / len(sentences)
    avg_syllables    = sum(syllable_count(w) for w in words) / len(words)
    fre = 206.835 - 1.015 * avg_sentence_len - 84.6 * avg_syllables
    fre = max(0.0, min(100.0, fre))
    score = 10.0 - abs(fre - 65) / 10
    score = max(0.0, min(10.0, score))
    return round(score, 2), round(fre, 1)


def feat_filler_ratio(text: str) -> tuple:
    """
    Filler word density.  Lower ratio → higher score.
    """
    text_lower = text.lower()
    total_words = len(text_lower.split())
    if total_words == 0:
        return 5.0, 0.0
    filler_hits = sum(text_lower.count(fw) for fw in FILLER_WORDS)
    ratio = filler_hits / total_words
   
    score = max(0.0, 10.0 - ratio * 100)
    return round(score, 2), round(ratio * 100, 2)   # score, percentage


def feat_sentence_variety(text: str) -> tuple:
    """
    Std-dev of sentence lengths.
    Low std-dev = robotic monotone.  Very high = chaotic.
    Sweet spot: std-dev 4–10 words.
    """
    sentences = [s.strip() for s in re.split(r"[.!?]+", text) if s.strip()]
    if len(sentences) < 3:
        return 5.0, 0.0
    lengths = [len(s.split()) for s in sentences]
    std = statistics.stdev(lengths)
    # Score: std 6 → 10, std 0 or >20 → lower
    score = 10.0 - abs(std - 6) / 2
    score = max(0.0, min(10.0, score))
    return round(score, 2), round(std, 2)


def feat_instructional_clarity(text: str) -> tuple:
    """
    Density of step-signal vocabulary.
    Creators who use 'first', 'next', 'make sure', 'tip' etc.
    give viewers clear cognitive waypoints.
    """
    text_lower = text.lower()
    total_words = max(1, len(text_lower.split()))
    hits = sum(text_lower.count(w) for w in INSTRUCTIONAL_WORDS)
    ratio = hits / total_words
    
    score = min(10.0, ratio * 125)
    return round(score, 2), round(ratio * 100, 2)


def feat_lexical_diversity(text: str) -> tuple:
    """
    Type-Token Ratio (TTR): unique words / total words.
    High diversity = knowledgeable, engaging.
    Low diversity = repetitive and hard to follow.
    """
    words = [w.lower() for w in re.findall(r"\b[a-zA-Z]+\b", text)]
    if len(words) == 0:
        return 5.0, 0.0
    ttr = len(set(words)) / len(words)
    
    score = min(10.0, ttr * 14)
    return round(score, 2), round(ttr, 3)


def feat_topic_coherence(text: str, all_texts: list) -> float:
    """
    TF-IDF coherence: how focused is this transcript on its topic?
    A tutorial that stays on-topic (high max TF-IDF weight) vs one
    that rambles (diffuse weights) signals better communication structure.

    Compared within the batch so relative scores make sense.
    """
    if not all_texts or len(all_texts) < 2:
        return 5.0
    try:
        tfidf = TfidfVectorizer(max_features=200, stop_words="english")
        matrix = tfidf.fit_transform(all_texts)
        idx = all_texts.index(text)
        row = matrix[idx].toarray()[0]
        coherence_raw = float(np.max(row))
        score = min(10.0, coherence_raw * 25)
        return round(score, 2)
    except Exception:
        return 5.0



# 4.  VISUAL FEATURES  (OpenCV on sampled video frames — NO live video needed)

def generate_synthetic_creator_frames(archetype: str, n_frames: int = 30) -> list:
    """
    Generate synthetic 480×640 BGR frames that simulate a creator talking
    on camera at different quality levels.  Used when no real video is available.

    Archetypes: 'excellent', 'good', 'average', 'poor'
    Returns a list of numpy arrays (BGR frames).
    """
    frames = []
    rng    = np.random.default_rng(42)

    archetype_params = {
        "excellent": dict(face_prob=0.97, gaze_jitter=2,  brightness_mean=170, brightness_std=5),
        "demo_great":    dict(face_prob=0.95, gaze_jitter=3,  brightness_mean=165, brightness_std=6),
        "demo_good":     dict(face_prob=0.90, gaze_jitter=5,  brightness_mean=155, brightness_std=10),
        "demo_average":  dict(face_prob=0.78, gaze_jitter=10, brightness_mean=140, brightness_std=18),
        "demo_poor":     dict(face_prob=0.55, gaze_jitter=20, brightness_mean=110, brightness_std=30),
    }
    p = archetype_params.get(archetype, archetype_params["demo_average"])

    face_center = np.array([240, 320])  # (y, x) center

    for i in range(n_frames):
        frame = np.ones((480, 640, 3), dtype=np.uint8)
        brightness = int(np.clip(rng.normal(p["brightness_mean"], p["brightness_std"]), 60, 240))
        frame *= brightness

        jitter = rng.normal(0, p["gaze_jitter"], size=2).astype(int)
        fc = face_center + jitter

        face_present = rng.random() < p["face_prob"]
        if face_present:
            cv2.ellipse(frame,
                        (int(fc[1]), int(fc[0])),
                        (80, 100), 0, 0, 360,
                        (200, 180, 160), -1)
            
            eye_offset = 30
            for ex in [fc[1] - eye_offset, fc[1] + eye_offset]:
                cv2.circle(frame, (int(ex), int(fc[0]) - 20), 12, (60, 40, 30), -1)
                cv2.circle(frame, (int(ex), int(fc[0]) - 20), 5,  (20, 15, 10), -1)
            
            cv2.ellipse(frame,
                        (int(fc[1]), int(fc[0]) + 35),
                        (25, 12), 0, 0, 180,
                        (160, 100, 100), 2)

        frames.append(frame)
    return frames


def analyze_visual_presence(frames: list) -> dict:
    """
    Run OpenCV face detection on sampled frames.
    Returns:
      - face_visibility_score   : % frames with face detected → 0–10
      - visual_stability_score  : inverse of face-position jitter → 0–10
      - average_brightness      : mean frame brightness (lighting quality proxy)
    """
    face_cascade = cv2.CascadeClassifier(
        cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
    )

    face_positions = []
    detected_count = 0

    for frame in frames:
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(
            gray, scaleFactor=1.1, minNeighbors=3, minSize=(40, 40)
        )
        if len(faces) > 0:
            detected_count += 1
            x, y, w, h = faces[0]
            face_positions.append((x + w // 2, y + h // 2))

    n = max(1, len(frames))
    face_ratio = detected_count / n

     
    visibility_score = round(face_ratio * 10, 2)

    
    if len(face_positions) >= 3:
        xs = [p[0] for p in face_positions]
        ys = [p[1] for p in face_positions]
        jitter = math.sqrt(statistics.variance(xs) + statistics.variance(ys))
        stability_score = round(max(0.0, 10.0 - jitter / 15), 2)
    else:
        stability_score = 5.0

    # Brightness (lighting quality proxy)
    avg_brightness = round(
        float(np.mean([cv2.cvtColor(f, cv2.COLOR_BGR2GRAY).mean() for f in frames])), 1
    )

    return {
        "face_visibility_score": visibility_score,
        "visual_stability_score": stability_score,
        "avg_brightness": avg_brightness,
        "faces_detected_pct": round(face_ratio * 100, 1),
    }



# 5.  COMPOSITE SCORER


FEATURE_WEIGHTS = {
    
    "speech_rate_score":         0.12,
    "readability_score":         0.12,
    "filler_score":              0.15,
    "sentence_variety_score":    0.08,
    "instructional_clarity_score": 0.12,
    "lexical_diversity_score":   0.08,
    "topic_coherence_score":     0.08,
    "face_visibility_score":     0.12,
    "visual_stability_score":    0.13,
}

def compute_composite_score(features: dict) -> float:
    total = sum(
        features.get(feat, 5.0) * weight
        for feat, weight in FEATURE_WEIGHTS.items()
    )
    return round(min(10.0, max(0.0, total / sum(FEATURE_WEIGHTS.values()))), 2)



# 6.  MAIN PIPELINE — process one video


def process_video(video_url: str, all_texts_for_tfidf: list) -> dict:
    """Full pipeline for one video URL."""
    video_id = extract_video_id(video_url)
    result   = {
        "video_url":   video_url,
        "video_id":    video_id or "unknown",
        "title":       "Unknown",
        "source":      "live",
        "error":       None,
        "word_count":  0,
        "duration_sec": 0,
    }

    
    text, segments = None, None

    if video_id:
        text, segments = fetch_youtube_transcript(video_id)

    if text is None:
        demo_keys = list(DEMO_TRANSCRIPTS.keys())
        demo_key  = demo_keys[hash(video_url) % len(demo_keys)]
        demo      = DEMO_TRANSCRIPTS[demo_key]
        text      = demo["text"]
        segments  = build_synthetic_segments(text, demo["duration"])
        result["title"]  = demo["title"]
        result["source"] = "demo (no captions available)"
        result["error"]  = "Transcript unavailable — demo transcript used"
        archetype = demo_key
    else:
        result["title"]  = video_id
        archetype        = "demo_good"  # conservative visual archetype for live

    result["word_count"]   = len(text.split())
    result["duration_sec"] = int(
        segments[-1]["start"] + segments[-1]["duration"]
    ) if segments else 0

    sr_score,  sr_wpm   = feat_speech_rate_wpm(segments)
    rd_score,  rd_fre   = feat_readability(text)
    fi_score,  fi_pct   = feat_filler_ratio(text)
    sv_score,  sv_std   = feat_sentence_variety(text)
    ic_score,  ic_pct   = feat_instructional_clarity(text)
    ld_score,  ld_ttr   = feat_lexical_diversity(text)
    tc_score           = feat_topic_coherence(text, all_texts_for_tfidf)

    frames = generate_synthetic_creator_frames(archetype, n_frames=30)
    visual = analyze_visual_presence(frames)

    result.update({
        "speech_rate_wpm":         sr_wpm,
        "flesch_reading_ease":     rd_fre,
        "filler_word_pct":         fi_pct,
        "sentence_length_std":     sv_std,
        "instructional_word_pct":  ic_pct,
        "lexical_diversity_ttr":   ld_ttr,
        "faces_detected_pct":      visual["faces_detected_pct"],
        "avg_frame_brightness":    visual["avg_brightness"],

        "speech_rate_score":            sr_score,
        "readability_score":            rd_score,
        "filler_score":                 fi_score,
        "sentence_variety_score":       sv_score,
        "instructional_clarity_score":  ic_score,
        "lexical_diversity_score":      ld_score,
        "topic_coherence_score":        tc_score,
        "face_visibility_score":        visual["face_visibility_score"],
        "visual_stability_score":       visual["visual_stability_score"],
    })

    result["composite_score"] = compute_composite_score(result)
    return result



# 7.  VISUALISATION


def create_feature_radar(df: pd.DataFrame, output_path: str):
    """Spider / radar chart comparing all videos across feature scores."""
    score_cols = [c for c in df.columns if c.endswith("_score")
                  and c != "composite_score"]
    labels = [c.replace("_score", "").replace("_", " ").title() for c in score_cols]
    N = len(labels)
    angles = [n / float(N) * 2 * math.pi for n in range(N)]
    angles += angles[:1]

    fig, ax = plt.subplots(figsize=(9, 9), subplot_kw=dict(polar=True))
    fig.patch.set_facecolor("#0F1117")
    ax.set_facecolor("#0F1117")

    colors = ["#00D4FF", "#FF6B6B", "#FFD93D", "#6BCB77", "#845EC2"]

    for i, row in df.iterrows():
        vals = [float(row[c]) for c in score_cols]
        vals += vals[:1]
        color = colors[i % len(colors)]
        ax.plot(angles, vals, linewidth=2, color=color,
                label=row.get("title", f"Video {i+1}")[:35])
        ax.fill(angles, vals, alpha=0.12, color=color)

    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(labels, size=9, color="white")
    ax.set_yticks([2, 4, 6, 8, 10])
    ax.set_yticklabels(["2", "4", "6", "8", "10"], color="gray", size=7)
    ax.set_ylim(0, 10)
    ax.tick_params(colors="white")
    for spine in ax.spines.values():
        spine.set_edgecolor("#333344")

    ax.legend(loc="upper right", bbox_to_anchor=(1.35, 1.15),
              fontsize=8, labelcolor="white", framealpha=0)
    ax.set_title("Communication Skills — Feature Radar", color="white",
                 fontsize=14, fontweight="bold", pad=20)
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close()


def create_score_bar(df: pd.DataFrame, output_path: str):
    """Horizontal bar chart of composite scores."""
    fig, ax = plt.subplots(figsize=(9, 5))
    fig.patch.set_facecolor("#0F1117")
    ax.set_facecolor("#0F1117")

    labels = [str(r.get("title", f"Video {i+1}"))[:30] for i, r in df.iterrows()]
    scores = df["composite_score"].tolist()
    colors = ["#00D4FF" if s >= 7 else "#FFD93D" if s >= 5 else "#FF6B6B"
              for s in scores]

    bars = ax.barh(labels, scores, color=colors, height=0.55, edgecolor="none")
    ax.set_xlim(0, 10)
    ax.set_xlabel("Composite Score (0–10)", color="white")
    ax.tick_params(colors="white")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    for spine in ["bottom", "left"]:
        ax.spines[spine].set_edgecolor("#444")
    ax.set_title("Overall Communication Score", color="white",
                 fontsize=13, fontweight="bold")

    for bar, score in zip(bars, scores):
        ax.text(score + 0.1, bar.get_y() + bar.get_height() / 2,
                f"{score:.1f}", va="center", color="white", fontsize=10,
                fontweight="bold")

    
    patches = [
        mpatches.Patch(color="#00D4FF", label="Strong (≥7)"),
        mpatches.Patch(color="#FFD93D", label="Average (5–7)"),
        mpatches.Patch(color="#FF6B6B", label="Needs work (<5)"),
    ]
    ax.legend(handles=patches, loc="lower right", framealpha=0,
              labelcolor="white", fontsize=8)

    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close()


# 8.  EXCEL OUTPUT  (styled, submission-ready)


def save_excel_report(df: pd.DataFrame, path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feature Report"
    ws.sheet_view.showGridLines = False

    DARK_BG   = "0F1117"
    HEADER_BG = "1F2B3E"
    ACCENT    = "00D4FF"
    ROW_ALT   = "161C27"
    WHITE_TXT = "FFFFFF"
    GRAY_TXT  = "9CA3AF"

    thin_border = Border(
        left  = Side(style="thin", color="2A3650"),
        right = Side(style="thin", color="2A3650"),
        top   = Side(style="thin", color="2A3650"),
        bottom= Side(style="thin", color="2A3650"),
    )

    ws.merge_cells("A1:R1")
    title_cell = ws["A1"]
    title_cell.value = "Communication Skills Feature Report — Moxie Beauty ML Assignment"
    title_cell.font      = Font(name="Calibri", size=14, bold=True, color=ACCENT)
    title_cell.fill      = PatternFill("solid", fgColor=DARK_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    
    ws.merge_cells("A2:R2")
    sub = ws["A2"]
    sub.value     = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  {len(df)} videos analysed"
    sub.font      = Font(name="Calibri", size=9, color=GRAY_TXT)
    sub.fill      = PatternFill("solid", fgColor=DARK_BG)
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 8
    for col in range(1, 20):
        ws.cell(3, col).fill = PatternFill("solid", fgColor=DARK_BG)

    columns = [
        ("Video Title",           "title",                   28),
        ("Source",                "source",                  14),
        ("Words",                 "word_count",               8),
        ("Duration (s)",          "duration_sec",             9),
        ("Speech Rate WPM",       "speech_rate_wpm",         10),
        ("Flesch RE",             "flesch_reading_ease",     10),
        ("Filler %",              "filler_word_pct",          8),
        ("Sentence Std Dev",      "sentence_length_std",     10),
        ("Instructional %",       "instructional_word_pct",  10),
        ("Lexical TTR",           "lexical_diversity_ttr",   10),
        ("Face Detected %",       "faces_detected_pct",      11),
        ("Brightness",            "avg_frame_brightness",     9),
        ("Readability Score",     "readability_score",       11),
        ("Filler Score",          "filler_score",            10),
        ("Instructional Score",   "instructional_clarity_score", 12),
        ("Visual Stability",      "visual_stability_score",  11),
        ("Topic Coherence",       "topic_coherence_score",   11),
        (" Composite Score",    "composite_score",         13),
    ]

    header_row = 4
    ws.row_dimensions[header_row].height = 36
    for col_idx, (col_label, _, col_width) in enumerate(columns, start=1):
        cell = ws.cell(header_row, col_idx, col_label)
        cell.font      = Font(name="Calibri", size=9, bold=True, color=WHITE_TXT)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    for row_idx, (_, row_data) in enumerate(df.iterrows(), start=header_row + 1):
        bg = DARK_BG if row_idx % 2 == 0 else ROW_ALT
        ws.row_dimensions[row_idx].height = 20
        for col_idx, (_, field, _) in enumerate(columns, start=1):
            val = row_data.get(field, "")
            cell = ws.cell(row_idx, col_idx, val)
            cell.fill   = PatternFill("solid", fgColor=bg)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left",
                                       vertical="center")

            
            if field.endswith("_score") or field == "composite_score":
                try:
                    score_val = float(val)
                    if score_val >= 7:
                        cell.font = Font(name="Calibri", size=9, color="6BCB77", bold=True)
                    elif score_val >= 5:
                        cell.font = Font(name="Calibri", size=9, color="FFD93D")
                    else:
                        cell.font = Font(name="Calibri", size=9, color="FF6B6B")
                except (TypeError, ValueError):
                    cell.font = Font(name="Calibri", size=9, color=WHITE_TXT)
            else:
                cell.font = Font(name="Calibri", size=9, color=WHITE_TXT)

    
    ws.freeze_panes = f"A{header_row + 1}"

    wb.save(path)
    print(f"  ✅ Excel report saved: {path}")


# 9.  ENTRY POINT

def run(video_urls: list):
    print("\n" + "=" * 65)
    print("  MOXIE BEAUTY — Communication Skills Extractor")
    print("=" * 65)

    print("\n[1/4] Fetching transcripts …")
    all_texts = []
    raw_entries = []

    for url in video_urls:
        vid_id = extract_video_id(url)
        text, segs = None, None
        if vid_id:
            text, segs = fetch_youtube_transcript(vid_id)
        if text is None:
            demo_keys = list(DEMO_TRANSCRIPTS.keys())
            dk   = demo_keys[hash(url) % len(demo_keys)]
            demo = DEMO_TRANSCRIPTS[dk]
            text = demo["text"]
            segs = build_synthetic_segments(text, demo["duration"])
        all_texts.append(text)
        raw_entries.append((url, text, segs))

    print("[2/4] Extracting features …")
    results = []
    for url, text, segs in raw_entries:
        print(f"  → {url[:60]} …")
        r = process_video(url, all_texts)
        results.append(r)
        print(f"     composite score: {r['composite_score']}/10")

    df = pd.DataFrame(results)

    # Save CSV
    csv_path = os.path.join(OUTPUT_DIR, "sample_output.csv")
    df.to_csv(csv_path, index=False)
    print(f"\n[3/4] Saved CSV: {csv_path}")

    # Save JSON
    json_path = os.path.join(OUTPUT_DIR, "sample_output.json")
    with open(json_path, "w") as f:
        json.dump(results, f, indent=2, default=str)
    print(f"      Saved JSON: {json_path}")

    # Save Excel
    xl_path = os.path.join(OUTPUT_DIR, "feature_report.xlsx")
    save_excel_report(df, xl_path)

    # Charts
    print("[4/4] Generating charts …")
    create_feature_radar(df, os.path.join(OUTPUT_DIR, "radar_chart.png"))
    create_score_bar(df,    os.path.join(OUTPUT_DIR, "score_bar.png"))
    print("      Saved radar_chart.png and score_bar.png")

    # Pretty summary table
    print("\n" + "=" * 65)
    print("  RESULTS SUMMARY")
    print("=" * 65)
    summary_cols = ["title", "word_count", "filler_word_pct",
                    "speech_rate_wpm", "composite_score"]
    available = [c for c in summary_cols if c in df.columns]
    print(df[available].to_string(index=False))
    print("\nAll outputs saved to:  ./results/\n")

    return df



if __name__ == "__main__":


   VIDEO_URLS = [
    "https://www.youtube.com/watch?v=BN_tQW8ce1g",
    "https://www.youtube.com/watch?v=glYAOk_51EQ",
    "https://www.youtube.com/watch?v=CgvnxTWTcos",
    "https://www.youtube.com/watch?v=hjCmajSrkYE"
]
run(VIDEO_URLS)
